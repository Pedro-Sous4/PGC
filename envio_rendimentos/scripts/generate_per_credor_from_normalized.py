# Generate per-credor files (PGC, EMISSAO, PRODUTIVIDADE, EXTRATO) from normalized sheets
import sys
import os
sys.path.append(r'C:\PGC\envio_rendimentos\envio_rendimentos')
sys.path.append(r'C:\PGC\envio_rendimentos')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'settings')
import django
django.setup()

import pandas as pd
import json
from core.utils_lgm import limpar_nome_credor

# Auto-detect PGC folders under MEDIA_ROOT/PGC
pgc_root = os.path.join('envio_rendimentos', 'arquivos_gerados', 'PGC')
# pick folder with the largest number of files (heuristic) or pass via arg
pgc_num = None
if len(sys.argv) > 1:
    pgc_num = sys.argv[1]
else:
    # prefer '34' if exists else pick the newest numeric folder
    candidate = os.path.join(pgc_root, '34')
    if os.path.exists(candidate):
        pgc_num = '34'
    else:
        # fallback - pick first numeric directory
        for name in sorted(os.listdir(pgc_root)):
            if name.isdigit():
                pgc_num = name
                break
if not pgc_num:
    raise SystemExit('Não foi possível detectar a pasta do PGC. Passe o número como argumento.')

pasta = os.path.join(pgc_root, str(int(pgc_num)))
print('Usando pasta PGC:', pasta)

# Helper utilities: tolerant filename and column detection
import unicodedata, re

def _normalize(s):
    if s is None:
        return ''
    s = str(s).lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'\W+', '_', s)
    return s

def find_file_with_keywords(directory, keywords):
    for fn in os.listdir(directory):
        name_norm = _normalize(fn)
        if all(k in name_norm for k in keywords):
            return os.path.join(directory, fn)
    return None

# Paths of normalized sheets (tolerant search)
path_base = find_file_with_keywords(pasta, ['base','pgc', str(int(pgc_num))]) or os.path.join(pasta, f'BASE PGC {pgc_num}.xlsx')
path_extrato = find_file_with_keywords(pasta, ['extrato']) or os.path.join(pasta, 'EXTRATO.xlsx')
path_prod = find_file_with_keywords(pasta, ['produtividad', 'produtiv', 'prod']) or os.path.join(pasta, 'PRODUTIVIDADE.xlsx')
path_emissao = find_file_with_keywords(pasta, ['emissao', 'emissao']) or os.path.join(pasta, 'EMISSAO.xlsx')
path_minimo = find_file_with_keywords(pasta, ['minimo']) or os.path.join(pasta, 'MINIMO.xlsx')

print('Arquivos detectados:')
print(' BASE:', path_base, os.path.exists(path_base))
print(' EXTRATO:', path_extrato, os.path.exists(path_extrato))
print(' PROD:', path_prod, os.path.exists(path_prod))
print(' EMISSAO:', path_emissao, os.path.exists(path_emissao))
print(' MINIMO:', path_minimo, os.path.exists(path_minimo))

# Read if exists (tolerant to read errors for large files like EXTRATO)
try:
    base_df = pd.read_excel(path_base) if os.path.exists(path_base) else pd.DataFrame()
except Exception as e:
    print('Erro lendo BASE:', e)
    base_df = pd.DataFrame()

try:
    extrato_df = pd.read_excel(path_extrato) if os.path.exists(path_extrato) else pd.DataFrame()
except Exception as e:
    print('Erro lendo EXTRATO (vai pular EXTRATO):', e)
    extrato_df = pd.DataFrame()

try:
    prod_df = pd.read_excel(path_prod) if os.path.exists(path_prod) else pd.DataFrame()
except Exception as e:
    print('Erro lendo PRODUTIVIDADE (vai pular):', e)
    prod_df = pd.DataFrame()

try:
    emissao_df = pd.read_excel(path_emissao) if os.path.exists(path_emissao) else pd.DataFrame()
except Exception as e:
    print('Erro lendo EMISSAO (vai derivar de BASE):', e)
    emissao_df = pd.DataFrame()

try:
    minimo_df = pd.read_excel(path_minimo) if os.path.exists(path_minimo) else pd.DataFrame()
except Exception as e:
    print('Erro lendo MINIMO (vai pular):', e)
    minimo_df = pd.DataFrame()

# Ensure 'credor' column exists by trying tolerant matches
def ensure_credor(df, df_name):
    if df.empty:
        return df
    cols = list(df.columns)
    norm_map = {c: _normalize(c) for c in cols}
    # direct match
    for c,n in norm_map.items():
        if n == 'credor' or 'credor' in n:
            if c != 'credor':
                df.rename(columns={c: 'credor'}, inplace=True)
            return df
    # fallback: look for nome/beneficiario
    for c,n in norm_map.items():
        if 'nome' in n or 'benef' in n or 'titular' in n:
            df.rename(columns={c: 'credor'}, inplace=True)
            print(f"Coluna '{c}' renomeada para 'credor' em {df_name}")
            return df
    return df

base_df = ensure_credor(base_df, 'BASE')
extrato_df = ensure_credor(extrato_df, 'EXTRATO')
prod_df = ensure_credor(prod_df, 'PRODUTIVIDADE')
emissao_df = ensure_credor(emissao_df, 'EMISSAO')
minimo_df = ensure_credor(minimo_df, 'MINIMO')

# Ensure credor column exists now
if 'credor' not in base_df.columns:
    raise SystemExit('Coluna "credor" não encontrada em BASE. Verifique a normalização.')

# Prepare credor entries and folder mapping
credores = base_df['credor'].dropna().unique()
credor_entries = []
for c in credores:
    display = str(c).strip()
    folder = os.path.join(pasta, display)
    os.makedirs(folder, exist_ok=True)
    credor_entries.append({'display': display, 'folder': folder, 'norm': _normalize(display)})

credor_map = {e['norm']: e for e in credor_entries}

# Helper: stream large EXTRATO into per-credor CSV files (openpyxl streaming)
import csv
import openpyxl

def stream_extrato_to_csv(path_extrato, credor_map):
    if not os.path.exists(path_extrato):
        return {}
    created = {}
    # Try openpyxl first (xlsx)
    try:
        wb = openpyxl.load_workbook(path_extrato, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            wb.close()
            return {}
        norm_headers = [_normalize(h) for h in header]
        credor_idx = None
        for i,n in enumerate(norm_headers):
            if 'credor' in n or 'nome' in n or 'benef' in n:
                credor_idx = i
                break
        if credor_idx is None:
            print('Coluna "credor" não encontrada em EXTRATO; pulando EXTRATO streaming')
            wb.close()
            return {}
        handles = {}
        writers = {}
        for row in it:
            if row is None:
                continue
            val = row[credor_idx]
            if val is None:
                continue
            norm = _normalize(val)
            entry = credor_map.get(norm)
            if not entry:
                continue
            if norm not in writers:
                out_fn = os.path.join(entry['folder'], f"{entry['display']} - EXTRATO {pgc_num}.csv")
                fh = open(out_fn, 'w', encoding='utf8', newline='')
                w = csv.writer(fh)
                w.writerow(header)
                handles[norm] = fh
                writers[norm] = w
                created[norm] = out_fn
            writers[norm].writerow([v for v in row])
        for fh in handles.values():
            fh.close()
        wb.close()
        return created
    except Exception as e:
        # Fallback: try reading as CSV in chunks
        print('Falha ao abrir EXTRATO com openpyxl (tentando como CSV):', e)
        tried_encodings = ['utf8','latin1']
        tried_seps = [None, ',', ';', '\t']
        for enc in tried_encodings:
            for sep in tried_seps:
                try:
                    for chunk in pd.read_csv(path_extrato, encoding=enc, sep=sep, engine='python', chunksize=10000):
                        # find credor column
                        cols = list(chunk.columns)
                        norm_map = {c: _normalize(c) for c in cols}
                        credor_col = None
                        for c,n in norm_map.items():
                            if 'credor' in n or 'nome' in n or 'benef' in n:
                                credor_col = c
                                break
                        if credor_col is None:
                            print('Coluna "credor" não encontrada em EXTRATO-CSV; pulando')
                            return {}
                        # group by credor
                        for name, grp in chunk.groupby(credor_col):
                            if pd.isna(name):
                                continue
                            norm = _normalize(name)
                            entry = credor_map.get(norm)
                            if not entry:
                                continue
                            out_fn = os.path.join(entry['folder'], f"{entry['display']} - EXTRATO {pgc_num}.csv")
                            if not os.path.exists(out_fn):
                                grp.to_csv(out_fn, index=False, encoding=enc)
                                created[norm] = out_fn
                            else:
                                grp.to_csv(out_fn, mode='a', header=False, index=False, encoding=enc)
                    return created
                except UnicodeDecodeError:
                    print(f'Falha decodificando EXTRATO com {enc}, tentando outro encoding...')
                    break
                except Exception as e2:
                    print('Falha ao processar EXTRATO com encoding', enc, 'sep', sep, ':', e2)
                    # try next separator
                    continue
        return {}

# Helper: stream produtividade similarly (if needed)
def stream_prod_to_csv(path_prod, credor_map):
    if not os.path.exists(path_prod):
        return {}
    created = {}
    try:
        wb = openpyxl.load_workbook(path_prod, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            wb.close()
            return {}
        norm_headers = [_normalize(h) for h in header]
        credor_idx = None
        for i,n in enumerate(norm_headers):
            if 'credor' in n or 'nome' in n or 'benef' in n:
                credor_idx = i
                break
        if credor_idx is None:
            print('Coluna "credor" não encontrada em PRODUTIVIDADE; pulando')
            wb.close()
            return {}
        handles = {}
        writers = {}
        for row in it:
            if row is None:
                continue
            val = row[credor_idx]
            if val is None:
                continue
            norm = _normalize(val)
            entry = credor_map.get(norm)
            if not entry:
                continue
            if norm not in writers:
                out_fn = os.path.join(entry['folder'], f"{entry['display']} - PRODUTIVIDADE {pgc_num}.csv")
                fh = open(out_fn, 'w', encoding='utf8', newline='')
                w = csv.writer(fh)
                w.writerow(header)
                handles[norm] = fh
                writers[norm] = w
                created[norm] = out_fn
            writers[norm].writerow([v for v in row])
        for fh in handles.values():
            fh.close()
        wb.close()
        return created
    except Exception as e:
        print('Falha abrindo PRODUTIVIDADE com openpyxl (tentando como CSV):', e)
        # fallback to CSV reading with encoding/sep attempts
        tried_encodings = ['utf8', 'latin1']
        tried_seps = [None, ',', ';', '\t']
        for enc in tried_encodings:
            for sep in tried_seps:
                try:
                    for chunk in pd.read_csv(path_prod, encoding=enc, sep=sep, engine='python', chunksize=10000):
                        cols = list(chunk.columns)
                        norm_map = {c: _normalize(c) for c in cols}
                        credor_col = None
                        for c, n in norm_map.items():
                            if 'credor' in n or 'nome' in n or 'benef' in n:
                                credor_col = c
                                break
                        if credor_col is None:
                            print('Coluna "credor" não encontrada em PRODUTIVIDADE-CSV; pulando')
                            return {}
                        for name, grp in chunk.groupby(credor_col):
                            if pd.isna(name):
                                continue
                            norm = _normalize(name)
                            entry = credor_map.get(norm)
                            if not entry:
                                continue
                            out_fn = os.path.join(entry['folder'], f"{entry['display']} - PRODUTIVIDADE {pgc_num}.csv")
                            if not os.path.exists(out_fn):
                                grp.to_csv(out_fn, index=False, encoding=enc)
                                created[norm] = out_fn
                            else:
                                grp.to_csv(out_fn, mode='a', header=False, index=False, encoding=enc)
                    return created
                except UnicodeDecodeError:
                    print(f'Falha decodificando PRODUTIVIDADE com {enc}, tentando outro encoding...')
                    break
                except Exception as e2:
                    print('Falha ao processar PRODUTIVIDADE com encoding', enc, 'sep', sep, ':', e2)
                    continue
        return {}

# Robust processing helpers for EXTRATO and PRODUTIVIDADE (tries multiple formats)

def robust_stream_extrato(path_extrato, credor_map):
    # Try existing xlsx handler first
    created = stream_extrato_to_csv(path_extrato, credor_map)
    if created:
        return created
    # Try old xls (OLE) via xlrd
    head = b''
    try:
        with open(path_extrato, 'rb') as fh:
            head = fh.read(8)
    except Exception:
        pass
    if head.startswith(b"\xD0\xCF"):
        try:
            import xlrd
            wb_x = xlrd.open_workbook(path_extrato, on_demand=True)
            sh = wb_x.sheet_by_index(0)
            if sh.nrows == 0:
                wb_x.release_resources()
                return {}
            header = sh.row_values(0)
            norm_headers = [_normalize(h) for h in header]
            credor_idx = None
            for i, n in enumerate(norm_headers):
                if 'credor' in n or 'nome' in n or 'benef' in n:
                    credor_idx = i
                    break
            if credor_idx is None:
                wb_x.release_resources()
                return {}
            created = {}
            handles = {}
            writers = {}
            for r in range(1, sh.nrows):
                row = sh.row_values(r)
                val = row[credor_idx]
                if val is None:
                    continue
                norm = _normalize(val)
                entry = credor_map.get(norm)
                if not entry:
                    continue
                if norm not in writers:
                    out_fn = os.path.join(entry['folder'], f"{entry['display']} - EXTRATO {pgc_num}.csv")
                    fh = open(out_fn, 'w', encoding='utf8', newline='')
                    w = csv.writer(fh)
                    w.writerow(header)
                    handles[norm] = fh
                    writers[norm] = w
                    created[norm] = out_fn
                writers[norm].writerow([v for v in row])
            for fh in handles.values():
                fh.close()
            wb_x.release_resources()
            return created
        except Exception as ex_xl:
            print('Falha ao processar EXTRATO como xls:', ex_xl)
    # Fallback CSV attempts with separators/encodings
    tried_encodings = ['utf8','latin1']
    tried_seps = [None, ',', ';', '\t']
    for enc in tried_encodings:
        for sep in tried_seps:
            try:
                for chunk in pd.read_csv(path_extrato, encoding=enc, sep=sep, engine='python', chunksize=10000):
                    cols = list(chunk.columns)
                    norm_map = {c: _normalize(c) for c in cols}
                    credor_col = None
                    for c,n in norm_map.items():
                        if 'credor' in n or 'nome' in n or 'benef' in n:
                            credor_col = c
                            break
                    if credor_col is None:
                        print('Coluna "credor" não encontrada em EXTRATO-CSV; pulando')
                        return {}
                    for name, grp in chunk.groupby(credor_col):
                        if pd.isna(name):
                            continue
                        norm = _normalize(name)
                        entry = credor_map.get(norm)
                        if not entry:
                            continue
                        out_fn = os.path.join(entry['folder'], f"{entry['display']} - EXTRATO {pgc_num}.csv")
                        if not os.path.exists(out_fn):
                            grp.to_csv(out_fn, index=False, encoding=enc)
                        else:
                            grp.to_csv(out_fn, mode='a', header=False, index=False, encoding=enc)
                return {}
            except UnicodeDecodeError:
                print(f'Falha decodificando EXTRATO com {enc}, tentando outro encoding...')
                break
            except Exception as e2:
                print('Falha ao processar EXTRATO com encoding', enc, 'sep', sep, ':', e2)
                continue
    return {}

# Stream EXTRATO and PRODUTIVIDADE into per-credor CSVs (if needed)
extrato_created = {}
prod_created = {}
if os.path.exists(path_extrato) and extrato_df.empty:
    print('Processando EXTRATO em streaming...')
    extrato_created = robust_stream_extrato(path_extrato, credor_map)
if os.path.exists(path_prod) and prod_df.empty:
    print('Processando PRODUTIVIDADE em streaming...')
    prod_created = stream_prod_to_csv(path_prod, credor_map)

# Build initial report structure
report = {'pgc': pgc_num, 'total_credores': int(len(credores)), 'credores': []}

for e in credor_entries:
    display = e['display']
    folder = e['folder']
    norm = e['norm']

    # Filter base rows for this credor
    mask = base_df['credor'].astype(str).str.strip() == display
    df_base_credor = base_df[mask]

    # PGC file
    pgc_name = f"{display} - PGC {pgc_num}.xlsx"
    path_pgc_cred = os.path.join(folder, pgc_name)
    df_base_credor.to_excel(path_pgc_cred, index=False)

    # EMISSAO: derive from base rows (columns present in emisao_df) or use emisao_df filtered
    if not emissao_df.empty:
        if 'credor' in emissao_df.columns:
            df_emissao_cred = emissao_df[emissao_df['credor'].astype(str).str.strip() == display]
        else:
            cols = list(emissao_df.columns)
            df_emissao_cred = df_base_credor[[c for c in df_base_credor.columns if c in cols]]
    else:
        df_emissao_cred = df_base_credor.copy()

    emissao_name = f"{display} - PGC {pgc_num} EMISSÃO.xlsx"
    path_emissao_cred = os.path.join(folder, emissao_name)
    df_emissao_cred.to_excel(path_emissao_cred, index=False)

    # PRODUTIVIDADE: prefer CSV created by streaming, else from df
    path_prod_cred = None
    if prod_created.get(norm):
        path_prod_cred = prod_created.get(norm)
    elif not prod_df.empty and 'credor' in prod_df.columns:
        df_prod_cred = prod_df[prod_df['credor'].astype(str).str.strip() == display]
        if not df_prod_cred.empty:
            prod_name = f"{display} - PRODUTIVIDADE {pgc_num}.xlsx"
            path_prod_cred = os.path.join(folder, prod_name)
            df_prod_cred.to_excel(path_prod_cred, index=False)

    # EXTRATO: prefer CSV created by streaming, else from df
    path_ext_cred = None
    if extrato_created.get(norm):
        path_ext_cred = extrato_created.get(norm)
    elif not extrato_df.empty and 'credor' in extrato_df.columns:
        df_ext_cred = extrato_df[extrato_df['credor'].astype(str).str.strip() == display]
        if not df_ext_cred.empty:
            ext_name = f"{display} - EXTRATO {pgc_num}.xlsx"
            path_ext_cred = os.path.join(folder, ext_name)
            df_ext_cred.to_excel(path_ext_cred, index=False)

    # MINIMO info (single-row lookup)
    minimo_info = None
    if not minimo_df.empty:
        rows = minimo_df[minimo_df['credor'].astype(str).str.strip().str.upper() == display.upper()]
        if not rows.empty:
            r = rows.iloc[0]
            minimo_info = {'valor': float(r['minimo']) if not pd.isna(r['minimo']) else None, 'empresa_emissao': r.get('empresa_emissao'), 'cnpj': r.get('cnpj')}

    report['credores'].append({
        'display': display,
        'files': {
            'pgc': path_pgc_cred,
            'emissao': path_emissao_cred,
            'produtividade': path_prod_cred,
            'extrato': path_ext_cred
        },
        'minimo': minimo_info
    })

# write report
out = os.path.join(pasta, f'per_credor_report_pgc_{pgc_num}.json')
with open(out, 'w', encoding='utf8') as f:
    json.dump(report, f, ensure_ascii=False, indent=2)

print('Geração por credor concluída. Report:', out)
print('Total credores:', len(report['credores']))
