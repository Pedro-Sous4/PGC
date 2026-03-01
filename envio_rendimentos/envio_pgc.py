import os
import smtplib
import openpyxl
from email.message import EmailMessage
from email.utils import formataddr

# -------------------------------
# CONFIGURAÇÕES DO E-MAIL (OUTLOOK)
# -------------------------------
EMAIL_HOST = "smtp.office365.com"
EMAIL_PORT = 587
EMAIL_HOST_USER = "contatopj@laghettogolden.com.br"
EMAIL_HOST_PASSWORD = "M#481199181005ud"  # <- senha de aplicativo, não a senha de login

NOME_ARQUIVO_PLANILHA = "Nomes e e-mails - Sports.xlsx"
PASTA_MACRO = r"C:\Users\ADMIN\Downloads\MACRO"

PGC_NUMERO = "PGC 14"
# -------------------------------


def enviar_email(destinatario, nome, anexos):
    msg = EmailMessage()

    msg["Subject"] = f"{PGC_NUMERO} - Envio de documentos"
    msg["From"] = formataddr(("Financeiro", EMAIL_HOST_USER))
    msg["To"] = destinatario

    corpo = f"""
Olá {nome},

No e-mail podem constar 3 planilhas, sendo elas:

- os valores de cada empresa para emissão - {PGC_NUMERO} EMISSÃO
- o borderô com as comissões que estão sendo pagas - {PGC_NUMERO}
- o histórico de comissões que foram pagas ou foram bloqueadas por inadimplência e/ou distrato - EXTRATO

AS NOTAS DEVEM SER EMITIDAS PARA AS EMPRESAS QUE CONSTAM NA PLANILHA "{PGC_NUMERO} EMISSÃO"

Notas devem ser enviadas até Domingo, dia 14/12.

Notas enviadas após o prazo, serão programadas para 15 dias após o recebimento.

Atenciosamente.
"""

    msg.set_content(corpo)

    # Anexar arquivos
    for arquivo in anexos:
        with open(arquivo, "rb") as f:
            dados = f.read()

        msg.add_attachment(
            dados,
            maintype="application",
            subtype="octet-stream",
            filename=os.path.basename(arquivo)
        )

    # Envio do e-mail
    with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
        server.starttls()
        server.login(EMAIL_HOST_USER, EMAIL_HOST_PASSWORD)
        server.send_message(msg)

    print(f"[OK] E-mail enviado para {nome} ({destinatario}).")


def buscar_arquivos_por_nome(nome):
    arquivos = []

    for arquivo in os.listdir(PASTA_MACRO):
        if arquivo.lower().startswith(nome.lower()):
            arquivos.append(os.path.join(PASTA_MACRO, arquivo))

    return arquivos


def enviar_para_um(nome_procurado, sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome, email = row[:2]

        if not nome:
            continue

        if nome_procurado.lower() in nome.lower():
            print(f"\nEncontrado: {nome} - {email}")

            anexos = buscar_arquivos_por_nome(nome)

            if not anexos:
                print(f"[ERRO] Nenhum arquivo encontrado para {nome}.")
                return

            print("\nArquivos encontrados:")
            for a in anexos:
                print(" -", os.path.basename(a))

            confirmar = input("\nDeseja enviar este e-mail? (s/n): ").strip().lower()

            if confirmar == "s":
                enviar_email(email, nome, anexos)
            else:
                print("Envio cancelado.")

            return

    print("\nNenhum nome correspondente encontrado.")


def enviar_para_todos(sheet):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nome, email = row[:2]

        if not nome or not email:
            continue

        print(f"\nProcessando: {nome} - {email}")

        anexos = buscar_arquivos_por_nome(nome)

        if not anexos:
            print(f"[ERRO] Nenhum arquivo encontrado para {nome}.")
            continue

        enviar_email(email, nome, anexos)


def main():
    print("Lendo planilha...")
    wb = openpyxl.load_workbook(NOME_ARQUIVO_PLANILHA)
    sheet = wb.active

    print("""
=============================
   SISTEMA DE ENVIO PGC
=============================
1 - Enviar para TODOS
2 - Enviar para APENAS UM
=============================
""")

    opcao = input("Escolha uma opção: ").strip()

    if opcao == "1":
        enviar_para_todos(sheet)

    elif opcao == "2":
        nome = input("Digite o nome ou parte do nome: ").strip()
        enviar_para_um(nome, sheet)

    else:
        print("Opção inválida. Encerrando.")


if __name__ == "__main__":
    main()
