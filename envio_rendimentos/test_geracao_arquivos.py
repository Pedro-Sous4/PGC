import os
import sys
import shutil
import pandas as pd
from datetime import datetime

# --- Configuração do Ambiente Django ---
# Adiciona o diretório do projeto ao sys.path para permitir importações
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(BASE_DIR)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "envio_rendimentos.settings")
import django
django.setup()
# --- Fim da Configuração ---

from django.conf import settings
from core.models import Credor
from core.utils import gerar_arquivos_credor, gerar_minimos_por_coluna_ap

# =============================================================================
# DADOS DE TESTE (MOCK)
# =============================================================================
NUMERO_PGC_TESTE = "99"
PERIODO_TESTE = "04/2025"
NOME_CREDOR_1 = "GUILHERME FOGASSA DA SILVA"
NOME_CREDOR_2 = "JOAO CARLOS PEREIRA"
NOME_CREDOR_3 = "MARIA (SEM MINIMO)" # Este não terá mínimo

def criar_planilha_pgc_falsa(caminho_completo):
    """Cria um arquivo Excel de PGC falso para os testes."""
    
    # --- Aba BASE PGC ---
    dados_base = {
        "Empresa": ["EMPRESA A", "EMPRESA B", "EMPRESA A"],
        "Credor": [NOME_CREDOR_1, NOME_CREDOR_1, NOME_CREDOR_2],
        "Documento": [101, 102, 103],
        "Cliente": ["Cliente A1", "Cliente B1", "Cliente A2"],
        "Parcela": ["1/3", "2/3", "1/1"],
        "Dt. emissão": [datetime(2025, 4, 1), datetime(2025, 4, 5), datetime(2025, 4, 8)],
        "Valor original": [1000.50, 500.00, 1200.75]
    }
    df_base = pd.DataFrame(dados_base)

    # --- Aba EXTRATO CREDOR ---
    dados_extrato = {
        "Empresa": ["EMPRESA A", "EMPRESA C"],
        "Credor": [NOME_CREDOR_1, NOME_CREDOR_2],
        "Documento": [101, 901],
        "Cliente": ["Cliente A1", "Cliente C1"],
        "Parcela": ["1/3", "1/2"],
        "Dt. emissão": [datetime(2025, 4, 1), datetime(2025, 3, 15)],
        "Valor original": [1000.50, 300.00],
        "Dt. vencimento": [datetime(2025, 5, 1), datetime(2025, 4, 15)],
        "Obs. baixa": ["", "INADIMPLENTE"]
    }
    df_extrato = pd.DataFrame(dados_extrato)

    # --- Aba PRODUTIVIDADE ---
    dados_prod = {
        "Empresa": ["EMPRESA B"],
        "Credor": [NOME_CREDOR_1],
        "Documento": [102],
        "Cliente": ["Cliente B1"],
        "Parcela": ["2/3"],
        "Dt. emissão": [datetime(2025, 4, 5)],
        "Valor original": [500.00],
        "Dt. vencimento": [datetime(2025, 6, 5)]
    }
    df_prod = pd.DataFrame(dados_prod)

    # --- Aba PGC (para extração de MÍNIMOS) ---
    # Criando um DataFrame com 50 colunas para simular a estrutura real
    colunas_pgc = [f'col_{i}' for i in range(50)]
    df_pgc = pd.DataFrame(columns=colunas_pgc)
    
    # Adicionando dados nas linhas e colunas corretas (AD, AJ, AK, AL, AP)
    # Lembre-se que os índices são baseados em 0
    # Linha 8 do Excel = índice 7 no DataFrame (após pular 7 linhas de cabeçalho)
    dados_pgc_minimos = [
        # Credor 1 (com mínimo)
        {29: NOME_CREDOR_1, 35: 1500, 36: "EMPRESA X", 37: "11.111.111/0001-11", 41: "EMPRESA X"},
        # Credor 2 (com mínimo)
        {29: NOME_CREDOR_2, 35: 2000, 36: "EMPRESA Y", 37: "22.222.222/0001-22", 41: "EMPRESA Y"},
        # Credor 3 (sem valor na coluna AP, não deve ir para o MINIMO.xlsx)
        {29: NOME_CREDOR_3, 35: 1800, 36: "EMPRESA Z", 37: "33.333.333/0001-33", 41: None},
    ]
    
    # Adiciona 7 linhas vazias para simular o cabeçalho
    df_pgc_final = pd.concat([pd.DataFrame(columns=colunas_pgc, index=range(7)), pd.DataFrame(dados_pgc_minimos)], ignore_index=True)
    # Renomeia as colunas para o teste
    df_pgc_final = df_pgc_final.rename(columns={
        29: "CREDOR", 
        35: "MINIMO/FIXO GARANTIDO PARA EMISSAO NF", 
        36: "EMPRESA EMISSAO NF", 
        37: "CNPJ",
        41: "EMPRESA_FLAG_AP" # Coluna AP
    })


    # --- Escreve o arquivo Excel ---
    with pd.ExcelWriter(caminho_completo, engine='openpyxl') as writer:
        df_base.to_excel(writer, sheet_name=f'BASE PGC {NUMERO_PGC_TESTE}', index=False)
        df_extrato.to_excel(writer, sheet_name='EXTRATO CREDOR', index=False)
        df_prod.to_excel(writer, sheet_name=f'PRODUTIVIDADE ABRIL-25', index=False)
        df_pgc_final.to_excel(writer, sheet_name=f'PGC {NUMERO_PGC_TESTE}', index=False)

    print(f"✅ Planilha de teste criada em: {caminho_completo}")
    # Ajusta células específicas na aba PGC para garantir colunas AD/AJ/AK/AL/AP preenchidas
    try:
        from openpyxl import load_workbook
        wb = load_workbook(caminho_completo)
        sheet_name = f'PGC {NUMERO_PGC_TESTE}'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # escreve os registros diretamente nas colunas corretas (0-based index +1 para openpyxl)
            for i, rec in enumerate(dados_pgc_minimos):
                row_idx = 8 + i  # começa na linha 8
                for key, val in rec.items():
                    col_idx = int(key) + 1
                    ws.cell(row=row_idx, column=col_idx, value=val)
            wb.save(caminho_completo)
    except Exception as e:
        print(f"⚠️ Aviso: não foi possível ajustar células PGC diretamente: {e}")


# =============================================================================
# FUNÇÕES DE VERIFICAÇÃO (ASSERT)
# =============================================================================
def verificar(condicao, msg_sucesso, msg_falha):
    """Função auxiliar para imprimir resultados do teste."""
    if condicao:
        print(f"  [SUCESSO] {msg_sucesso}")
        return True
    else:
        print(f"  [FALHA]   {msg_falha}")
        return False

# =============================================================================
# SCRIPT DE TESTE PRINCIPAL
# =============================================================================
def rodar_teste():
    """Executa o teste de ponta a ponta."""
    
    # --- 1. SETUP DO AMBIENTE DE TESTE ---
    print("\n--- INICIANDO TESTE DE GERAÇÃO DE ARQUIVOS ---\n")
    pasta_base_pgc = os.path.join(settings.MEDIA_ROOT, 'PGC', NUMERO_PGC_TESTE)
    
    # Limpa a pasta de teste anterior, se existir
    if os.path.exists(pasta_base_pgc):
        shutil.rmtree(pasta_base_pgc)
        print(f"🧹 Pasta de teste antiga removida: {pasta_base_pgc}")
        
    os.makedirs(pasta_base_pgc)
    print(f"📂 Pasta de teste criada: {pasta_base_pgc}")

    caminho_planilha_falsa = os.path.join(pasta_base_pgc, f'PGC {NUMERO_PGC_TESTE}.xlsx')
    criar_planilha_pgc_falsa(caminho_planilha_falsa)
    
    # --- 2. EXECUÇÃO DA LÓGICA DE GERAÇÃO ---
    
    # a) Garante arquivo EMPRESAS_NOMECURTO_CNPJ.xlsx para lookup de CNPJ
    emp_path = r"C:\PGC\envio_rendimentos\arquivos_gerados\EMPRESAS_NOMECURTO_CNPJ.xlsx"
    os.makedirs(os.path.dirname(emp_path), exist_ok=True)
    emp_df = pd.DataFrame({'nome_curto': ['EMPRESA X', 'EMPRESA Y'], 'cnpj': ['11.111.111/0001-11', '22.222.222/0001-22']})
    emp_df.to_excel(emp_path, index=False)

    # b) Gerar arquivo de MÍNIMOS
    print("\n▶️ Executando geração do arquivo MINIMO.xlsx...")
    caminho_minimo_gerado = gerar_minimos_por_coluna_ap(caminho_planilha_falsa, NUMERO_PGC_TESTE, pasta_saida=pasta_base_pgc)

    # b) Ler os DataFrames das abas para passar para a função de gerar arquivos por credor
    planilhas_dict = pd.read_excel(caminho_planilha_falsa, sheet_name=None)
    base_df = planilhas_dict.get(f'BASE PGC {NUMERO_PGC_TESTE}')
    extrato_df = planilhas_dict.get('EXTRATO CREDOR')
    prod_df = planilhas_dict.get('PRODUTIVIDADE ABRIL-25')
    # Normaliza colunas para formato esperado pelas funções (lowercase, underscores)
    from core.utils import normalizar_colunas_simples
    if base_df is not None:
        base_df = normalizar_colunas_simples(base_df)
    if extrato_df is not None:
        extrato_df = normalizar_colunas_simples(extrato_df)
    if prod_df is not None:
        prod_df = normalizar_colunas_simples(prod_df)
    minimo_df = pd.read_excel(caminho_minimo_gerado) if caminho_minimo_gerado and os.path.exists(caminho_minimo_gerado) else None

    # c) Gerar arquivos para cada credor
    credores_para_testar = [NOME_CREDOR_1, NOME_CREDOR_2]
    for nome_credor in credores_para_testar:
        print(f"\n▶️ Executando geração de arquivos para o credor: {nome_credor}...")
        # Cria ou obtém o Credor de forma resiliente usando o helper centralizado
        credor_obj, _ = Credor.get_or_create_by_nome(nome_credor, defaults={'periodo': PERIODO_TESTE})
        # Garantir que o período esteja preenchido para gerar o nome da produtividade
        try:
            credor_obj.periodo = PERIODO_TESTE
            credor_obj.save(update_fields=['periodo'])
        except Exception:
            pass
        
        gerar_arquivos_credor(
            credor=credor_obj,
            numero_pgc=NUMERO_PGC_TESTE,
            base_df=base_df,
            extrato_df=extrato_df,
            prod_df=prod_df,
            minimo_df=minimo_df,
            pasta_pgc=pasta_base_pgc
        )

    # --- 3. VERIFICAÇÃO DOS RESULTADOS ---
    print("\n--- VERIFICANDO RESULTADOS ---\n")
    sucesso_total = True

    # a) Verificar arquivo MINIMO.xlsx
    print("1. Verificando arquivo MINIMO.xlsx:")
    caminho_minimo_final = os.path.join(pasta_base_pgc, 'MINIMO.xlsx')
    if not verificar(os.path.exists(caminho_minimo_final), "Arquivo MINIMO.xlsx existe na raiz do PGC.", "Arquivo MINIMO.xlsx NÃO foi encontrado na raiz do PGC."):
        sucesso_total = False
    else:
        df_min_leitura = pd.read_excel(caminho_minimo_final)
        if not verificar(len(df_min_leitura) == 2, f"Arquivo contém 2 credores com mínimo (esperado: 2, encontrado: {len(df_min_leitura)}).", f"Número de credores incorreto."):
            sucesso_total = False
        if not verificar(NOME_CREDOR_3.upper() not in df_min_leitura['credor'].str.upper().values, "Credor 'MARIA (SEM MINIMO)' não está no arquivo.", "Credor sem mínimo foi incluído indevidamente."):
            sucesso_total = False

        # Verifica se CNPJs foram obtidos da planilha EMPRESAS_NOMECURTO_CNPJ.xlsx
        cnpjs_encontrados = set(str(x).strip() for x in df_min_leitura['cnpj'].dropna().unique())
        cnpjs_esperados = {'11.111.111/0001-11', '22.222.222/0001-22'}
        if not verificar(cnpjs_esperados.issubset(cnpjs_encontrados), f"CNPJs esperados encontrados: {cnpjs_esperados}", f"CNPJs não batem. Encontrados: {cnpjs_encontrados}"):
            sucesso_total = False

    # b) Verificar estrutura de pastas e arquivos por credor
    for nome_credor in credores_para_testar:
        print(f"\n2. Verificando pasta e arquivos para: {nome_credor}")
        
        # O nome da pasta deve ser MAIÚSCULO e com espaços
        pasta_credor_esperada = os.path.join(pasta_base_pgc, nome_credor.upper())
        
        if not verificar(os.path.isdir(pasta_credor_esperada), f"Pasta '{nome_credor.upper()}' existe.", f"Pasta '{nome_credor.upper()}' NÃO existe ou não é um diretório."):
            sucesso_total = False
            continue # Pula para o próximo credor se a pasta não existe

        # Lista de arquivos esperados
        arquivos_esperados = [
            f"{nome_credor.upper()} - PGC {str(NUMERO_PGC_TESTE).zfill(3)}.xlsx",
            f"{nome_credor.upper()} - PGC {str(NUMERO_PGC_TESTE).zfill(3)} EMISSÃO.xlsx",
            f"{nome_credor.upper()} - EXTRATO.xlsx",
            f"{nome_credor.upper()} - PRODUTIVIDADE ABRIL 2025.xlsx",
        ]
        
        arquivos_encontrados = os.listdir(pasta_credor_esperada)
        
        for arq_esperado in arquivos_esperados:
            # O arquivo de produtividade pode não ser gerado se não houver dados
            if "PRODUTIVIDADE" in arq_esperado and nome_credor == NOME_CREDOR_2:
                if not verificar(arq_esperado not in arquivos_encontrados, "Arquivo de PRODUTIVIDADE (corretamente) não foi gerado para credor sem dados.", "Arquivo de PRODUTIVIDADE foi gerado indevidamente."):
                    sucesso_total = False
                continue

            if not verificar(arq_esperado in arquivos_encontrados, f"Arquivo '{arq_esperado}' existe.", f"Arquivo '{arq_esperado}' NÃO foi encontrado."):
                sucesso_total = False

        if not verificar(len(arquivos_encontrados) <= len(arquivos_esperados), f"Não há arquivos extras na pasta (encontrados: {len(arquivos_encontrados)}).", f"Foram encontrados arquivos extras na pasta: {arquivos_encontrados}"):
            sucesso_total = False


    # --- 4. CONCLUSÃO ---
    print("\n--- CONCLUSÃO DO TESTE ---")
    if sucesso_total:
        print("\n✅✅✅ SUCESSO! Todos os testes passaram. A estrutura de arquivos e pastas está correta.")
    else:
        print("\n❌❌❌ FALHA! Um ou mais testes falharam. Revise as mensagens de erro acima.")
    
    # print(f"\nPara inspecionar os resultados manualmente, verifique a pasta: {pasta_base_pgc}")


if __name__ == "__main__":
    rodar_teste()
