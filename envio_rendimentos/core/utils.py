import os
import re
import unicodedata
import logging
import tempfile
from datetime import datetime
import pandas as pd
from django.core.mail import EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from weasyprint import HTML
from email.header import Header
from .models import Credor, HistoricoPGC, EmpresaPagadora
from difflib import get_close_matches
import openpyxl
from datetime import datetime
import logging
logger = logging.getLogger(__name__)



mes_atual = datetime.today().strftime('%m/%Y')
# Pega o dia 16 do mês atual
hoje = datetime.today()
dia_16 = datetime(hoje.year, hoje.month, 16)

# Nome do dia da semana em português (ex: "segunda-feira")
dias_semana = ['segunda-feira', 'terça-feira', 'quarta-feira', 'quinta-feira', 'sexta-feira', 'sábado', 'domingo']
nome_dia_semana = dias_semana[dia_16.weekday()]


# Configuração de logger
logger = logging.getLogger("envios")
logger.setLevel(logging.DEBUG)
if not logger.handlers:
    handler = logging.FileHandler(os.path.join(settings.MEDIA_ROOT, 'envios.log'))
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)


def normalizar_nome(nome):
    if not nome:
        return ''
    nome = re.sub(r'^\d+\s*-\s*', '', str(nome))
    nome = re.sub(r'\s*\([^)]*\)', '', nome)
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('ASCII')
    nome = re.sub(r'\s+', ' ', nome).strip().upper()
    return str(nome).strip().upper().replace("  ", " ")

def salvar_planilha_temporaria(file, numero_pgc):
    pasta = os.path.join(settings.MEDIA_ROOT, 'TEMPORARIOS')
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, f'PGC_{numero_pgc}_ORIGINAL.xlsx')
    with open(caminho, 'wb+') as destino:
        for chunk in file.chunks():
            destino.write(chunk)
    return caminho

def normalizar_planilha_origem(file_path, numero_pgc):
    renomear = {
        'Dt. emissão': 'dt_emissao',
        'Dt. vencimento': 'dt_vencimento',
        'Dt. baixa': 'dt_baixa',
        'Obs. baixa': 'obs_baixa'
    }
    df_dict = pd.read_excel(file_path, sheet_name=None)
    planilhas_tratadas = {
        aba: df.rename(columns=lambda col: renomear.get(str(col).strip(), str(col).strip()))
        for aba, df in df_dict.items()
    }
    pasta_saida = os.path.join('media', 'planilhas_originais_tratadas')
    os.makedirs(pasta_saida, exist_ok=True)
    caminho_final = os.path.join(pasta_saida, f'PGC {numero_pgc}.xlsx')
    with pd.ExcelWriter(caminho_final, engine='openpyxl') as writer:
        for aba, df in planilhas_tratadas.items():
            df.to_excel(writer, sheet_name=aba, index=False)
    return caminho_final

def normalizar_colunas_com_duas_linhas(df, header_start=5):
    # Força conversão para string antes do join
    df.columns = (
        df.iloc[header_start:header_start+2]
        .fillna('')
        .astype(str)  # <- transforma tudo em string
        .agg(' '.join)
        .str.strip()
        .str.lower()
        .str.replace(' ', '_')
        .str.replace('.', '')
    )
    df = df.iloc[header_start+2:].reset_index(drop=True)
    return df

def salvar_minimos_como_excel(df_minimos, numero_pgc):
    pasta = os.path.join(settings.MEDIA_ROOT, 'PGC', str(numero_pgc))
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, 'MINIMO.xlsx')
    df_minimos.to_excel(caminho, index=False)
    return caminho

def extrair_dados_planilhas(planilhas_dict, numero_pgc):
    base_df = produtividade_df = extrato_df = aba_pgcs = None
    for nome_aba, df in planilhas_dict.items():
        nome = nome_aba.strip().lower()
        if 'base' in nome:
            base_df = normalizar_colunas_simples(df.copy())
        elif 'produtividade' in nome:
            produtividade_df = normalizar_colunas_simples(df.copy())
        elif 'extrato' in nome:
            extrato_df = normalizar_colunas_simples(df.copy())
        elif nome.startswith(f"pgc {str(numero_pgc).lower()}"):
            aba_pgcs = df.copy()
    if base_df is None:
        raise ValueError('A aba "BASE" não foi encontrada.')
    return base_df, produtividade_df, extrato_df, aba_pgcs

# ===============================
# MENSAGENS DE E-MAIL PADRÃO
# ===============================

import json
import os

MENSAGEM_PADRAO = """{credor.nome},

Olá,

Segue em anexo produtividade, relatório com os bloqueios de comissão (distrato e inadimplência) e relação de clientes repassados.

No e-mail constam 4 planilhas, sendo elas:
- Os valores de cada empresa para emissão - PGC {historico.numero_pgc} EMISSÃO
- O borderô com os clientes que estão sendo repassados - PGC {historico.numero_pgc}
- A produtividade que está com o nome PRODUTIVIDADE {historico.periodo}
- O histórico das comissões bloqueadas por inadimplência e/ou distrato - EXTRATO

{info_minimo}
{info_descontos}


Notas devem ser enviadas até SEXTA-FEIRA, dia 16/{historico.periodo}, às 12:00h.

Informamos que o endereço da empresa ALTOS DA BORGES EMPREENDIMENTOS IMOBILIÁRIOS LTDA foi alterado para o seguinte local:
Rua Luiz de Camões, 360, Vila Nova, Novo Hamburgo/RS – CEP: 93.520-280.

Ressaltamos que, a partir desta data, não serão aceitas notas fiscais emitidas com o endereço antigo.

Atenciosamente,
"""

INFO_MINIMO_PADRAO = """Mínimo garantido no valor de {valor_formatado}. Emitir nota para {empresa} - {cnpj}."""
INFO_DESCONTOS_PADRAO = """Desconto de {valor_formatado} na empresa {empresa}, referente a {tipo}."""
# Texto padrão para a seção de descontos (editable pelo usuário)
INFO_DESCONTOS_PADRAO = """Desconto de {valor_formatado} na empresa {empresa}, referente a {tipo}."""

def _mensagens_path():
    """Retorna o caminho do arquivo onde as mensagens personalizadas são salvas."""
    from django.conf import settings
    pasta = os.path.join(settings.MEDIA_ROOT, 'mensagens')
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, 'mensagens_email.json')


def carregar_mensagens():
    """Carrega as mensagens salvas, ou retorna as padrões se o arquivo não existir."""
    caminho = _mensagens_path()
    if os.path.exists(caminho):
        with open(caminho, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        'mensagem': MENSAGEM_PADRAO,
        'info_minimo': INFO_MINIMO_PADRAO,
        'info_descontos': INFO_DESCONTOS_PADRAO
    }


def salvar_mensagens(mensagem, info_minimo, info_descontos=None):
    """Salva as mensagens personalizadas em JSON."""
    caminho = _mensagens_path()
    dados = {
        'mensagem': mensagem,
        'info_minimo': info_minimo,
        'info_descontos': info_descontos or INFO_DESCONTOS_PADRAO
    }
    with open(caminho, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

def _normalize_colname(c):
    import unicodedata, re
    if c is None:
        return ''
    s = str(c).strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = s.encode('ascii', 'ignore').decode('utf-8')
    s = re.sub(r'[^a-z0-9]+', '_', s)
    s = re.sub(r'_+', '_', s).strip('_')
    return s


from .minimo_display import format_empresa_para_exibicao


#TESTE
def gerar_arquivos_credor(credor, numero_pgc, base_df, extrato_df=None, prod_df=None, minimo_df=None, pasta_pgc=None):
    def nome_limpo(texto):
        texto = re.sub(r"^\d+\s*-\s*", "", str(texto))
        texto = re.sub(r"\s*\([^)]*\)", "", texto)
        texto = unicodedata.normalize('NFKD', texto.upper())
        return ''.join(c for c in texto if not unicodedata.combining(c)).strip()

    def formatar_nomes_colunas(df):
        """Formata nomes de colunas em MAIÚSCULAS sem abreviações"""
        mapa_colunas = {
            'empresa': 'EMPRESA',
            'credor': 'CREDOR',
            'documento': 'DOCUMENTO',
            'cliente': 'CLIENTE',
            'parcela': 'PARCELA',
            'dt_emissao': 'DATA EMISSÃO',
            'dt_vencimento': 'DATA VENCIMENTO',
            'valor_original': 'VALOR',
            'obs_baixa': 'OBSERVAÇÕES BAIXA',
            'cnpj': 'CNPJ',
            'cnpj para emissão': 'CNPJ PARA EMISSÃO'
        }
        
        novo_mapa = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            novo_mapa[col] = mapa_colunas.get(col_lower, col.upper())
        
        return df.rename(columns=novo_mapa)
    
    def ajustar_largura_colunas(workbook):
        """Ajusta largura das colunas automaticamente baseado no conteúdo"""
        for worksheet in workbook.sheetnames:
            ws = workbook[worksheet]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value:
                            # Calcula comprimento do conteúdo
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Define largura mínima de 12 e máxima de 50
                adjusted_width = min(max(max_length + 2, 12), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

    nome_credor_normalizado = nome_limpo(credor.nome)
    # nome_para_exibicao: usar o nome do credor para exibição (MAIÚSCULAS, com espaços)
    nome_para_exibicao = str(credor.nome).strip().upper()

    # Se pasta_pgc não informada, tentar inferir a partir do número do PGC
    if not pasta_pgc:
        try:
            pasta_pgc = os.path.join(settings.MEDIA_ROOT, 'PGC', str(int(numero_pgc)))
        except Exception:
            pasta_pgc = os.path.join(settings.MEDIA_ROOT, 'PGC', str(numero_pgc))

    pasta_origem = pasta_pgc

    # Pastas de credores devem usar o NOME PARA EXIBIÇÃO (sem underscores)
    pasta_saida = os.path.join(pasta_origem, f'{nome_para_exibicao}')
    os.makedirs(pasta_saida, exist_ok=True)

    def carregar_df(nome_arquivo):
        caminho = os.path.join(pasta_origem, nome_arquivo)
        return pd.read_excel(caminho) if os.path.exists(caminho) else None

    # Caminho do arquivo de mínimos (padronizado para MINIMO.xlsx)
    minimo_path = os.path.join(pasta_origem, 'MINIMO.xlsx')
    # Caso não exista com o nome exato, tenta variações case-insensitive
    if not os.path.exists(minimo_path):
        for f in os.listdir(pasta_origem):
            if f.lower() in ('minimo.xlsx', 'mínimo.xlsx'):
                minimo_path = os.path.join(pasta_origem, f)
                break

    # Se ainda não existir, tenta gerar o MINIMO.xlsx lendo a aba PGC {numero_pgc} (verificando coluna AP)
    if not os.path.exists(minimo_path):
        # tenta localizar o arquivo PGC na pasta (ex: 'PGC 26.xlsx' ou 'PGC 026.xlsx')
        arquivo_pgc = None
        for f in os.listdir(pasta_origem):
            if f.upper().startswith('PGC') and str(numero_pgc) in f:
                arquivo_pgc = os.path.join(pasta_origem, f)
                break
        if arquivo_pgc:
            try:
                gerado = gerar_minimos_por_coluna_ap(arquivo_pgc, numero_pgc, pasta_saida=pasta_origem)
                if gerado:
                    minimo_path = gerado
            except Exception as e:
                logger.warning(f"[MÍNIMO] Falha ao tentar gerar MINIMO.xlsx automaticamente: {e}")

    arquivos = {}

    # Garantir que os dataframes são cópias e que possuem as colunas
    # 'credor' e 'credor_normalizado' antes de qualquer processamento.
    for label in ('base', 'extrato', 'prod'):
        df = {'base': base_df, 'extrato': extrato_df, 'prod': prod_df}[label]
        if df is None:
            continue
        df = df.copy()

        # tenta renomear coluna semelhante para 'credor'
        if 'credor' not in df.columns:
            renamed = False
            for col in df.columns:
                if 'credor' in str(col).lower():
                    df = df.rename(columns={col: 'credor'})
                    renamed = True
                    break

        # se ainda não houver coluna 'credor', preenche com o nome do credor
        if 'credor' not in df.columns:
            df.loc[:, 'credor'] = credor.nome

        # garante coluna normalizada
        df.loc[:, 'credor_normalizado'] = df['credor'].astype(str).apply(nome_limpo)

        # reatribui ao nome original
        if label == 'base':
            base_df = df
        elif label == 'extrato':
            extrato_df = df
        else:
            prod_df = df

    numero_pgc_str = str(numero_pgc).zfill(3)

    # === BASE
    if base_df is not None:
        base_df['credor_normalizado'] = base_df['credor'].astype(str).apply(nome_limpo)
        df_base = base_df[base_df['credor_normalizado'] == nome_credor_normalizado]
        colunas_base = ['empresa', 'credor', 'documento', 'cliente', 'parcela', 'dt_emissao', 'valor_original']
        if not df_base.empty and all(col in df_base.columns for col in colunas_base):
            df_formatado = formatar_nomes_colunas(df_base[colunas_base].copy())
            arquivos[f'{nome_para_exibicao} - PGC {numero_pgc_str}.xlsx'] = df_formatado

        # === EMISSÃO
        # Apenas gera emissão se houver coluna 'empresa'
        if not df_base.empty and 'empresa' in df_base.columns:
            emissao_rows = []
            CAMINHO_EMPRESAS = r"C:\PGC\envio_rendimentos\arquivos_gerados\EMPRESAS_NOMECURTO_CNPJ.xlsx"

            try:
                df_empresas = pd.read_excel(CAMINHO_EMPRESAS)
                df_empresas['empresa_normalizada'] = df_empresas['nome_curto'].astype(str).apply(nome_limpo)
            except Exception:
                df_empresas = pd.DataFrame()

            for empresa, grupo in df_base.groupby('empresa'):
                empresa_limpa = nome_limpo(empresa)
                cnpj = None

                if not df_empresas.empty:
                    linha = df_empresas[df_empresas['empresa_normalizada'] == empresa_limpa]
                    if not linha.empty:
                        cnpj = linha.iloc[0]['cnpj']

                cnpj = cnpj if cnpj else "CNPJ NÃO ENCONTRADO"

                emissao_rows.append({
                    'EMPRESA': empresa,
                    'CREDOR': credor.nome,
                    'CNPJ PARA EMISSÃO': cnpj,
                    'VALOR': grupo['valor_original'].sum()
                })

            df_emissao = pd.DataFrame(emissao_rows)
            arquivos[f'{nome_para_exibicao} - PGC {numero_pgc_str} EMISSÃO.xlsx'] = df_emissao

    # === EXTRATO
    if extrato_df is not None:
        extrato_df['credor_normalizado'] = extrato_df['credor'].astype(str).apply(nome_limpo)
        df_ext = extrato_df[extrato_df['credor_normalizado'] == nome_credor_normalizado]
        colunas_ext = ['empresa', 'credor', 'documento', 'cliente', 'parcela', 'dt_emissao', 'valor_original', 'dt_vencimento']
        if not df_ext.empty and all(col in df_ext.columns for col in colunas_ext):
            final_cols = colunas_ext + (['obs_baixa'] if 'obs_baixa' in df_ext.columns else [])
            df_formatado = formatar_nomes_colunas(df_ext[final_cols].copy())
            arquivos[f'{nome_para_exibicao} - EXTRATO.xlsx'] = df_formatado

    # === PRODUTIVIDADE
    if prod_df is not None:
        prod_df['credor_normalizado'] = prod_df['credor'].astype(str).apply(nome_limpo)
        df_prod = prod_df[prod_df['credor_normalizado'] == nome_credor_normalizado]
        # torna dt_vencimento opcional: exige as colunas essenciais e adiciona dt_vencimento se existir
        colunas_prod_required = ['empresa', 'credor', 'documento', 'cliente', 'parcela', 'dt_emissao', 'valor_original']
        optional_cols = ['dt_vencimento']
        if not df_prod.empty and all(col in df_prod.columns for col in colunas_prod_required):
            import locale
            from calendar import month_name

            try:
                # Configura localidade brasileira
                locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')  # Linux/macOS
            except:
                locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')  # Windows

            try:
                mes, ano = map(int, credor.periodo.split('/'))
                data = pd.to_datetime(f"{ano}-{mes:02d}")
                mes_ano = data.strftime("%B %Y").upper()  # Ex: "ABRIL 2025"
            except:
                mes_ano = datetime.today().strftime("%B %Y").upper()  # fallback

            final_cols = colunas_prod_required + [c for c in optional_cols if c in df_prod.columns]
            df_formatado = formatar_nomes_colunas(df_prod[final_cols].copy())
            arquivos[f'{nome_para_exibicao} - PRODUTIVIDADE {mes_ano}.xlsx'] = df_formatado

    # NOTE: O arquivo MINIMO deve ser criado apenas no nível do PGC
    # A geração automática de MINIMO.xlsx (se necessário) já ocorre acima
    # e **não** será criada por credor aqui.

    # === SALVAR
    from .formatting import format_workbook
    for nome_arquivo, df in arquivos.items():
        caminho_final = os.path.join(pasta_saida, nome_arquivo)
        df.to_excel(caminho_final, index=False)

        # Apply consistent workbook formatting (headers, widths, number formats)
        try:
            format_workbook(caminho_final)
        except Exception as e:
            logger.warning(f"Não foi possível aplicar formatação em {nome_arquivo}: {e}")


#logger = logging.getLogger(__name__)

def encontrar_coluna_semelhante(coluna_alvo, colunas_existentes):
    correspondencias = get_close_matches(coluna_alvo.lower(), colunas_existentes, n=1, cutoff=0.6)
    return correspondencias[0] if correspondencias else None



'''###############################################################'''

def gerar_pdf_relatorio(credor):
    html_string = render_to_string('core/relatorio_pdf.html', {'Credor': credor})
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as output:
        HTML(string=html_string).write_pdf(output.name)
        return output.name

def enviar_email_com_arquivos(credor):
    import locale
    from .utils import carregar_mensagens
    mensagens = carregar_mensagens()
    mensagem_template = mensagens.get("mensagem", MENSAGEM_PADRAO)
    info_minimo_template = mensagens.get("info_minimo", INFO_MINIMO_PADRAO)
    info_descontos_template = mensagens.get("info_descontos", INFO_DESCONTOS_PADRAO)

    historico = credor.historicos.order_by('-data_envio').first()
    if not credor.email:
        logger.error(f'Credor {credor.nome} não possui e-mail cadastrado.')
        return False
    if not historico:
        logger.error(f'Credor {credor.nome} não possui histórico PGC registrado.')
        return False
    # >>> COLE ISTO AQUI <<<
    numero_pgc = str(historico.numero_pgc).zfill(3)

    base_pgc = os.path.join(settings.MEDIA_ROOT, 'PGC', numero_pgc)

    pasta_credor = encontrar_pasta_case_insensitive(
        base_pgc,
        credor.nome
    )

    if not pasta_credor:
        logger.error(
            f'Pasta não encontrada para {credor.nome}. '
            f'Base: {base_pgc}'
        )
        return False

    arquivos = [os.path.join(pasta_credor, f) for f in os.listdir(pasta_credor) if f.endswith('.xlsx')]
    if not arquivos:
        logger.error(f'Nenhum arquivo gerado para {credor.nome}.')
        return False

    # === Busca o mínimo garantido usando função centralizada
    info_minimo = ''
    info_minimo_dict = obter_minimo_garantido_para_credor(credor.nome, str(int(historico.numero_pgc) if str(historico.numero_pgc).isdigit() else historico.numero_pgc))

    if info_minimo_dict:
        try:
            valor = float(info_minimo_dict.get('valor', 0) or 0)
        except Exception:
            valor = 0.0
        # Define localidade para formato brasileiro
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')  # Linux/macOS
        except:
            try:
                locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')  # Windows
            except Exception:
                pass

        try:
            valor_formatado = locale.currency(valor, grouping=True)
        except Exception:
            # Fallback manual formatting
            valor_formatado = f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        empresa = info_minimo_dict.get('empresa', '')
        cnpj = info_minimo_dict.get('cnpj', '')

        # === Ajuste APENAS para exibição no e-mail: remover prefixo numérico "{numero} - " se presente
        empresa_exibicao = format_empresa_para_exibicao(empresa)

        info_minimo = info_minimo_template.format(valor_formatado=valor_formatado, empresa=empresa_exibicao, cnpj=cnpj)
        logger.info(f"[MINIMO] MÍNIMO ENCONTRADO → valor={valor_formatado}, empresa={empresa} (exib: {empresa_exibicao}), cnpj={cnpj}")
    else:
        logger.info(f"[MINIMO] Nenhum mínimo encontrado para {credor.nome} em PGC {historico.numero_pgc}")

    # === Monta info_descontos para o e-mail
    try:
        from .utils import formatar_info_descontos_para_email
        info_descontos = formatar_info_descontos_para_email(credor.nome, historico.numero_pgc, template=info_descontos_template)
        if info_descontos:
            logger.info(f"[DESCONTOS] Informação de descontos preparada para {credor.nome}")
        else:
            logger.info(f"[DESCONTOS] Nenhuma informação de descontos para {credor.nome}")
    except Exception as e:
        logger.warning(f"[DESCONTOS] Falha ao preparar info_descontos: {e}")
        info_descontos = ''

    # === Corpo do e-mail
    assunto = f"PGC {historico.numero_pgc}"
#     mensagem = f"""{credor.nome},

# Olá,

# Segue em anexo produtividade, relatório com os bloqueios de comissão (distrato e inadimplência) e relação de clientes repassados.

# No e-mail constam 4 planilhas, sendo elas:
# - Os valores de cada empresa para emissão - PGC {historico.numero_pgc} EMISSÃO
# - o borderô com os clientes que estão sendo repassados - PGC {historico.numero_pgc}
# - a produtividade que está com o nome PRODUTIVIDADE {historico.periodo}
# - o histórico das comissões que ficaram bloqueadas por inadimplência e/ou distrato - EXTRATO

# A PARTIR DE SETEMBRO/2024 AS NOTAS DEVEM SER EMITIDAS PARA AS EMPRESAS QUE CONSTAM NA PLANILHA "PGC {historico.numero_pgc} EMISSÃO".

# {info_minimo}
# Notas devem ser enviadas até QUARTA-FEIRA, dia 15/{mes_atual}.
# Notas enviadas após o prazo serão programadas para 15 dias após o recebimento.

# Atenciosamente,
# """
    def _render_mensagem_segura(template_str, credor_obj, historico_obj, info_minimo_str, info_descontos_str, mes_atual_str, nome_dia_str):
        # Substituições diretas para placeholders conhecidos, evitando KeyError em str.format
        try:
            s = str(template_str)
        except Exception:
            s = ''

        # valores simples
        credor_nome = getattr(credor_obj, 'nome', str(credor_obj)) if credor_obj is not None else ''
        historico_num = getattr(historico_obj, 'numero_pgc', str(historico_obj)) if historico_obj is not None else ''
        historico_periodo = getattr(historico_obj, 'periodo', '') if historico_obj is not None else ''

        replacements = {
            '{credor.nome}': str(credor_nome),
            '{historico.numero_pgc}': str(historico_num),
            '{historico.periodo}': str(historico_periodo),
            '{info_minimo}': str(info_minimo_str or ''),
            '{info_descontos}': str(info_descontos_str or ''),
            '{mes_atual}': str(mes_atual_str or ''),
            '{nome_dia_semana}': str(nome_dia_str or ''),
        }

        for k, v in replacements.items():
            s = s.replace(k, v)

        # Assegura que chaves globais com formato {key} também sejam substituídas quando presente
        try:
            # preencher chaves restantes de forma segura usando format_map com default vazio
            class _D(dict):
                def __missing__(self, key):
                    return ''
            s = s.format_map(_D(**{
                'info_minimo': info_minimo_str or '',
                'info_descontos': info_descontos_str or '',
                'mes_atual': mes_atual_str or '',
                'nome_dia_semana': nome_dia_str or ''
            }))
        except Exception:
            # se falhar, retorna a string parcialmente substituída
            pass

        return s

    mensagem = _render_mensagem_segura(mensagem_template, credor, historico, info_minimo, info_descontos, mes_atual, nome_dia_semana)

    # === Envio (Notas devem ser enviadas até às 12h de {nome_dia_semana.upper()}, dia 16/{mes_atual}.))
    # Encode subject explicitly as UTF-8 to avoid header mis-decodings downstream
    try:
        subject_encoded = str(Header(assunto, 'utf-8'))
    except Exception:
        subject_encoded = assunto
    email = EmailMessage(subject_encoded, mensagem, settings.DEFAULT_FROM_EMAIL, [credor.email])
    try:
        email.encoding = 'utf-8'
    except Exception:
        pass
    for arq in arquivos:
        email.attach_file(arq)

    # Registra tentativa de envio no EmailLog
    try:
        from .models import EmailLog
        from django.utils import timezone as dj_timezone
        log, created = EmailLog.objects.get_or_create(historico=historico, credor=credor, defaults={
            'numero_pgc': historico.numero_pgc or 0,
            'status': 'sending',
            'attempts': 1,
            'last_attempt_at': dj_timezone.now(),
        })
        if not created:
            log.status = 'sending'
            log.attempts = (log.attempts or 0) + 1
            log.last_attempt_at = dj_timezone.now()
            log.save()
    except Exception:
        log = None

    try:
        email.send()
        logger.info(f"E-mail enviado com sucesso para {credor.nome} ({credor.email}) com {len(arquivos)} arquivos.")
        # Atualiza credor
        try:
            credor.enviado = True
            credor.data_envio = dj_timezone.now()
            credor.save(update_fields=['enviado', 'data_envio'])
        except Exception:
            pass
        # Atualiza EmailLog
        try:
            if log:
                log.status = 'sent'
                log.sent_at = dj_timezone.now()
                log.error_message = None
                log.save()
        except Exception:
            pass
        return True
    except Exception as e:
        logger.error(f"Erro ao enviar e-mail para {credor.nome}: {e}")
        # marca credor como não enviado
        try:
            credor.enviado = False
            credor.save(update_fields=['enviado'])
        except Exception:
            pass
        # atualiza EmailLog com erro
        try:
            if log:
                log.status = 'failed'
                log.error_message = str(e)
                log.save()
        except Exception:
            pass
        return False

def normalizar_colunas_simples(df):
    df.columns = (
        df.columns
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace('.', '', regex=False)
        .str.replace(' ', '_')
        .str.replace('ã', 'a')
        .str.replace('é', 'e')
        .str.replace('ç', 'c')
        .str.replace('ê', 'e')
        .str.replace('í', 'i')
    )
    return df

def normalizar_e_salvar_planilha_base(path_origem, numero_pgc, pgc_prefix=None):
    import os
    import pandas as pd
    from django.conf import settings

    planilhas = pd.read_excel(path_origem, sheet_name=None)

    # =============================
    # PASTA ÚNICA DO PGC (usar sempre o número como nome da pasta)
    # Mantemos o `pgc_prefix` como metadado para agrupar via campo `Grupo`,
    # mas a pasta em disco será padronizada como o número do PGC (sem zeros à esquerda).
    # Isso evita inconsistências em outras partes do sistema que já
    # esperam `MEDIA_ROOT/PGC/<numero_pgc>`.
    try:
        numero_pasta = str(int(numero_pgc))
    except Exception:
        numero_pasta = str(numero_pgc)

    pasta_pgc = os.path.join(
        settings.MEDIA_ROOT,
        "PGC",
        numero_pasta
    )

    os.makedirs(pasta_pgc, exist_ok=True)

    numero_pgc_str = str(numero_pgc).zfill(2)
    pgc_tag = f"pgc{numero_pgc_str}"

    for nome_aba, df in planilhas.items():
        nome = nome_aba.strip().lower().replace(" ", "").replace("_", "")

        if nome.startswith(f"base{pgc_tag}"):
            df_base = normalizar_colunas_simples(df)
            df_base.to_excel(
                os.path.join(pasta_pgc, f"BASE PGC {numero_pgc}.xlsx"),
                index=False
            )

        elif any(k in nome for k in ["extrato", "exrato"]):
            df_ext = normalizar_colunas_simples(df)
            df_ext.to_excel(
                os.path.join(pasta_pgc, "EXTRATO.xlsx"),
                index=False
            )

        elif "produtividade" in nome or "liderança" in nome or "carteira" in nome or "vendas" in nome:
            df_prod = normalizar_colunas_simples(df)
            df_prod.to_excel(
                os.path.join(pasta_pgc, "PRODUTIVIDADE.xlsx"),
                index=False
            )

        elif nome == pgc_tag:
            df.to_excel(
                os.path.join(pasta_pgc, f"PGC {numero_pgc}.xlsx"),
                index=False
            )

    return pasta_pgc

from core.normalizacao import normalizar_nome_completo

def _progress_file(folder):
    return os.path.join(folder, "progress.json")

def write_progress(folder, data):
    os.makedirs(folder, exist_ok=True)
    with open(_progress_file(folder), "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def get_progress(request_id):
    from django.conf import settings
    path = os.path.join(
        settings.MEDIA_ROOT,
        "processing",
        request_id,
        "progress.json"
    )

    if not os.path.exists(path):
        return None

    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
    

    import os

def ensure_dir(path):
    """
    Garante que o diretório exista e retorna o caminho.
    """
    os.makedirs(path, exist_ok=True)
    return path

def encontrar_pasta_case_insensitive(pasta_base, nome_credor):
    """
    Localiza a pasta real do credor dentro do PGC,
    ignorando:
    - maiúsculas/minúsculas
    - acentos
    - espaços extras
    """

    if not os.path.isdir(pasta_base):
        logger.error(f"[PASTA] Base do PGC não existe: {pasta_base}")
        return None

    nome_norm = normalizar_nome_completo(nome_credor)

    for pasta in os.listdir(pasta_base):
        caminho = os.path.join(pasta_base, pasta)

        if not os.path.isdir(caminho):
            continue

        pasta_norm = normalizar_nome_completo(pasta)

        if pasta_norm == nome_norm:
            logger.info(f"[PASTA] Match encontrado: {pasta}")
            return caminho

    # Note: do not search inside a 'CREDORES' subfolder; creditor folders must live
    # directly under the PGC folder. If teams still have a legacy 'CREDORES' directory,
    # run the migration script `scripts/fix_credores_structure.py` to move folders up.

    logger.error(
        f"[PASTA] Nenhuma pasta encontrada para '{nome_credor}' "
        f"dentro de {pasta_base}"
    )
    return None




def obter_minimo_garantido_para_credor(nome_credor, numero_pgc):
    """Procura o mínimo garantido dentro da pasta do PGC.

    * Accepts either 'MINIMO.xlsx' or 'minimo.xlsx' (case-insensitive)
    * Normaliza nomes de colunas para localizar: credor, minimo, empresa, cnpj
    """
    pgc_dir = os.path.join(settings.MEDIA_ROOT, 'PGC', str(numero_pgc))
    logger.info(f"[MINIMO] Procurando arquivo MINIMO em: {pgc_dir}")
    if not os.path.isdir(pgc_dir):
        logger.warning(f"[MINIMO] Pasta PGC não encontrada: {pgc_dir}")
        return None

    # Find any file with name starting with 'minimo' (case-insensitive)
    candidato = None
    for fname in os.listdir(pgc_dir):
        if fname.lower().startswith('minimo') and fname.lower().endswith(('.xlsx', '.xls')):
            candidato = os.path.join(pgc_dir, fname)
            break

    if not candidato:
        logger.warning(f"[MINIMO] Arquivo MINIMO.* não encontrado em {pgc_dir}")
        return None

    logger.info(f"[MINIMO] Usando arquivo: {candidato}")

    try:
        df = pd.read_excel(candidato)
        logger.info(f"[MINIMO] Colunas detectadas: {list(df.columns)}")

        # Build normalized column map
        col_map = { _normalize_colname(c): c for c in df.columns }

        # find candidate columns
        col_credor = None
        col_minimo = None
        col_empresa = None
        col_cnpj = None

        for key, orig in col_map.items():
            if 'credor' in key or 'consultor' in key or 'corretor' in key:
                col_credor = orig
            if 'minim' in key or 'minimo' in key or 'fixo' in key:
                col_minimo = orig
            if 'empresa' in key and ('emiss' in key or 'emissa' in key or 'emit' in key or 'empresa' in key):
                col_empresa = orig
            if 'cnpj' in key:
                col_cnpj = orig

        # Fallback to more permissive matching
        if not col_minimo:
            for key, orig in col_map.items():
                if 'valor' in key and 'min' in key:
                    col_minimo = orig
        if not col_empresa:
            for key, orig in col_map.items():
                if 'empresa' in key:
                    col_empresa = orig

        if not col_credor or not col_minimo:
            logger.warning(f"[MINIMO] Colunas obrigatórias não encontradas (credor/minimo). Colunas: {list(df.columns)}")
            return None

        nome_normalizado = normalizar_nome(nome_credor)
        for _, row in df.iterrows():
            val_credor = str(row[col_credor]) if col_credor in row and not pd.isna(row[col_credor]) else ''
            if normalizar_nome(val_credor) == nome_normalizado:
                empresa_val = row[col_empresa] if col_empresa and col_empresa in row and not pd.isna(row[col_empresa]) else ''
                cnpj_val = row[col_cnpj] if col_cnpj and col_cnpj in row and not pd.isna(row[col_cnpj]) else ''
                minimo_val = row[col_minimo] if col_minimo in row and not pd.isna(row[col_minimo]) else None
                return {
                    'valor': minimo_val,
                    'empresa': empresa_val,
                    'cnpj': cnpj_val
                }
    except Exception as e:
        logger.error(f"[MÍNIMO] Erro ao ler mínimo.xlsx: {e}")
    return None


def _localizar_arquivo_descontos(numero_pgc):
    """Procura por um arquivo DESCONTOS.xlsx no diretório do PGC ou na pasta arquivos_gerados como fallback."""
    candidatos = []
    # prioridade: MEDIA_ROOT/PGC/<numero_pgc>/DESCONTOS.xlsx
    try:
        media_path = os.path.join(settings.MEDIA_ROOT, 'PGC', str(int(numero_pgc) if str(numero_pgc).isdigit() else numero_pgc))
    except Exception:
        media_path = os.path.join(settings.MEDIA_ROOT, 'PGC', str(numero_pgc))

    candidatos.append(os.path.join(media_path, 'DESCONTOS.xlsx'))

    # fallback: project arquivos_gerados path (where utils_lgm grava)
    base_proj = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'arquivos_gerados', 'PGC', str(int(numero_pgc) if str(numero_pgc).isdigit() else numero_pgc))
    candidatos.append(os.path.join(base_proj, 'DESCONTOS.xlsx'))

    for c in candidatos:
        if os.path.exists(c):
            logger.info(f"[DESCONTOS] Arquivo localizado: {c}")
            return c
    logger.info(f"[DESCONTOS] Nenhum arquivo DESCONTOS.xlsx encontrado para PGC {numero_pgc}")
    return None


def obter_descontos_para_credor(nome_credor, numero_pgc):
    """Retorna lista de descontos (dicionários) encontrados para o credor no DESCONTOS.xlsx.

    Cada item tem chaves: 'credor', 'valor', 'empresa', 'tipo' (quando disponível).
    """
    caminho = _localizar_arquivo_descontos(numero_pgc)
    if not caminho:
        return []

    try:
        df = pd.read_excel(caminho)
    except Exception as e:
        logger.warning(f"[DESCONTOS] Falha ao ler {caminho}: {e}")
        return []

    if df.empty:
        return []

    # Normalizar colunas para facilitar busca
    cols_map = {c.lower(): c for c in df.columns}
    # Possíveis nomes de coluna: 'credor', 'CREDOR', 'valor', 'VALOR', 'empresa', 'EMPRESA_DESCONTO', 'tipo'
    cred_col = cols_map.get('credor') or cols_map.get('credi') or next((v for k, v in cols_map.items() if 'credor' in k), None)
    val_col = cols_map.get('valor') or next((v for k, v in cols_map.items() if 'valor' in k), None)
    emp_col = cols_map.get('empresa_desconto') or cols_map.get('empresa') or next((v for k, v in cols_map.items() if 'empresa' in k), None)
    tipo_col = cols_map.get('tipo') or next((v for k, v in cols_map.items() if 'tipo' in k), None)

    if not cred_col or not val_col:
        logger.info(f"[DESCONTOS] Colunas esperadas não encontradas em {caminho}: {list(df.columns)}")
        return []

    resultados = []
    nome_norm = normalizar_nome(nome_credor)
    for _, row in df.iterrows():
        try:
            row_cred = str(row[cred_col]) if cred_col in row and not pd.isna(row[cred_col]) else ''
            if not row_cred:
                continue
            if normalizar_nome(row_cred) != nome_norm:
                continue

            valor = row[val_col] if val_col in row and not pd.isna(row[val_col]) else None
            empresa = row[emp_col] if emp_col and emp_col in row and not pd.isna(row[emp_col]) else ''
            tipo = row[tipo_col] if tipo_col and tipo_col in row and not pd.isna(row[tipo_col]) else ''

            resultados.append({
                'credor': row_cred,
                'valor': valor,
                'empresa': empresa,
                'tipo': tipo
            })
        except Exception:
            continue

    logger.info(f"[DESCONTOS] {len(resultados)} descontos encontrados para {nome_credor} em PGC {numero_pgc}")
    return resultados


def formatar_info_descontos_para_email(nome_credor, numero_pgc, template=None):
    """Gera o texto de descontos para inclusão no e-mail usando o template fornecido."""
    registros = obter_descontos_para_credor(nome_credor, numero_pgc)
    if not registros:
        return ''

    # define template
    info_template = template or INFO_DESCONTOS_PADRAO

    # Se o template não contém placeholders esperados, assumimos que o usuário
    # forneceu o texto final (pré-formatado) no card — nesse caso, não devemos
    # aplicar o template uma vez por registro (isso causa duplicação).
    placeholders = ('{valor_formatado}', '{empresa}', '{tipo}')
    if not any(ph in info_template for ph in placeholders):
        return str(info_template).strip()

    # preparar formatação monetária
    try:
        import locale
        try:
            locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
        except Exception:
            try:
                locale.setlocale(locale.LC_ALL, 'Portuguese_Brazil.1252')
            except Exception:
                pass
    except Exception:
        pass

    linhas = []
    for r in registros:
        try:
            valor = float(r.get('valor') or 0)
        except Exception:
            valor = 0.0

        # formata valor em BRL
        try:
            import locale
            valor_formatado = locale.currency(valor, grouping=True)
        except Exception:
            valor_formatado = f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')

        empresa = r.get('empresa') or ''
        tipo = r.get('tipo') or ''

        linhas.append(info_template.format(valor_formatado=valor_formatado, empresa=empresa, tipo=tipo))

    return '\n'.join(linhas)

def extrair_minimos_com_base_em_titulos(df):
    colunas_esperadas = {
        'credor': 'credor',
        'minimofixo_garantido_para_emissao_nf': 'minimo',
        'empresa_emissao_nf': 'empresa',
        'cnpj': 'cnpj'
    }
    colunas_existentes = [col.lower() for col in df.columns]
    mapeamento = {}

    for alvo, novo_nome in colunas_esperadas.items():
        coluna_encontrada = encontrar_coluna_semelhante(alvo, colunas_existentes)
        if not coluna_encontrada:
            raise ValueError(f'Coluna semelhante a "{alvo}" não encontrada.')
        mapeamento[coluna_encontrada] = novo_nome

    df = df.rename(columns=mapeamento)
    return df[['credor', 'minimo', 'empresa', 'cnpj']].dropna(subset=['credor'])

def extrair_minimos_por_coluna_fixa(caminho_arquivo, numero_pgc):
    wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    aba_nome = f"PGC {numero_pgc}"
    aba = wb[aba_nome] if aba_nome in wb.sheetnames else wb.active
    ws = aba

    dados = []

    for row in ws.iter_rows(min_row=8):
        if len(row) < 43:
            continue

        try:
            credor = row[32].value      # AG
            minimo = row[40].value      # AO
            empresa = row[41].value     # AP
            # AQ (42) é CNPJ da PGC — IGNORAR
        except IndexError:
            continue

        if credor and minimo and empresa:
            dados.append({
                'credor': str(credor).strip(),
                'minimo': minimo,
                'empresa': str(empresa).strip(),
                'cnpj': None  # CNPJ será buscado externamente
            })

    if not dados:
        raise ValueError("Nenhuma linha válida encontrada na aba de mínimos.")

    return pd.DataFrame(dados)

def gerar_minimos_por_coluna_ap(caminho_arquivo, numero_pgc, pasta_saida=None):
    """
    Gera um arquivo MINIMO.xlsx a partir da aba PGC {numero_pgc}
    verificando a coluna AP (index 41) a partir da linha 8.

    NOVA ESTRUTURA PGC:
        CREDOR            -> AG (index 32)
        MINIMO            -> AO (index 40)
        EMPRESA EMISSÃO   -> AP (index 41)
        CNPJ (PGC)        -> AQ (index 42)

    Retorna:
        Caminho do MINIMO.xlsx gerado ou None se nada encontrado.
    """

    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    except Exception as e:
        logger.warning(f"[MÍNIMO] Não foi possível abrir arquivo PGC para extrair mínimos: {e}")
        return None

    aba_nome = f"PGC {numero_pgc}"
    aba = wb[aba_nome] if aba_nome in wb.sheetnames else wb.active
    ws = aba

    CAMINHO_EMPRESAS = r"C:\PGC\envio_rendimentos\arquivos_gerados\EMPRESAS_NOMECURTO_CNPJ.xlsx"

    def _normalize_key(s):
        import unicodedata, re
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = unicodedata.normalize('NFKD', s)
        s = s.encode('ascii', 'ignore').decode('utf-8')
        s = re.sub(r"[^a-z0-9]+", ' ', s).strip()
        return s

    registros = []

    # ==============================
    # Carrega mapa EMPRESA -> CNPJ
    # ==============================
    try:
        df_empresas = pd.read_excel(CAMINHO_EMPRESAS)
        df_empresas.columns = [str(c).strip() for c in df_empresas.columns]

        col_nome = None
        col_cnpj = None

        for c in df_empresas.columns:
            lc = c.lower()
            if 'nome' in lc and ('curto' in lc or 'curt' in lc):
                col_nome = c
            if 'cnpj' in lc:
                col_cnpj = c

        if col_nome and col_cnpj:
            empresas_map = {
                _normalize_key(r[col_nome]): str(r[col_cnpj]).strip()
                for _, r in df_empresas.iterrows()
            }
        else:
            empresas_map = {}

    except Exception:
        empresas_map = {}

    # ==============================
    # Percorre linhas do PGC
    # ==============================
    for row in ws.iter_rows(min_row=8):

        try:
            empresa_emissao = row[41].value if len(row) > 41 else None  # AP
        except Exception:
            empresa_emissao = None

        if empresa_emissao and str(empresa_emissao).strip():

            credor = row[32].value if len(row) > 32 else None  # AG
            minimo = row[40].value if len(row) > 40 else None  # AO
            cnpj_pgc = row[42].value if len(row) > 42 else None  # AQ

            # ==========================
            # PRIORIDADE: CNPJ DO PGC
            # ==========================
            cnpj_final = None

            if cnpj_pgc and str(cnpj_pgc).strip():
                cnpj_final = str(cnpj_pgc).strip()
            else:
                # fallback para lookup por nome da empresa
                key = _normalize_key(empresa_emissao)
                cnpj_lookup = empresas_map.get(key, '')
                cnpj_final = cnpj_lookup if cnpj_lookup else None

            registros.append({
                'credor': str(credor).strip() if credor else None,
                'minimo': minimo,
                'empresa': str(empresa_emissao).strip(),
                'cnpj': cnpj_final
            })

    if not registros:
        logger.info(f"[MÍNIMO] Nenhum registro com AP preenchido encontrado na aba PGC {numero_pgc}.")
        return None

    df_minimos = pd.DataFrame(registros)

    if not pasta_saida:
        pasta_saida = os.path.dirname(caminho_arquivo)

    caminho_minimo = os.path.join(pasta_saida, 'MINIMO.xlsx')

    try:
        df_minimos.to_excel(caminho_minimo, index=False)
        logger.info(f"[MÍNIMO] Arquivo MINIMO.xlsx gerado em: {caminho_minimo}")
        return caminho_minimo
    except Exception as e:
        logger.error(f"[MÍNIMO] Falha ao salvar MINIMO.xlsx: {e}")
        return None


def processar_minimo_e_descontos_unificado(caminho_arquivo, numero_pgc, pasta_saida=None, salvar_minimo=True, salvar_descontos=False):
    """
    Percorre a aba PGC {numero_pgc} APENAS UMA VEZ (ws.iter_rows(min_row=8))
    e monta três listas separadas:
      - registros_minimos: [{'credor','minimo','empresa','cnpj'}, ...]
      - registros_retencao: [{'credor','valor','empresa'}, ...]
      - registros_cestas: [{'credor','valor','empresa'}, ...]

    Regras principais:
      - MÍNIMO: considera somente linhas com AP (index 41) preenchido. CNPJ é obtido
        via lookup em EMPRESAS_NOMECURTO_CNPJ.xlsx (mesma lógica usada atualmente).
      - DESCONTOS: se AH (index 33) tiver valor -> registro de retenção (empresa em AI index34).
                  se AJ (index 35) tiver valor -> registro de cesta (empresa em AK index36).

    Compatibilidade:
      - Não altera outras funções; serve como versão otimizada que evita múltiplos loops.
      - Mantém logs semelhantes aos já existentes para mínimos.
    """

    try:
        wb = openpyxl.load_workbook(caminho_arquivo, data_only=True)
    except Exception as e:
        logger.warning(f"[MÍNIMO] Não foi possível abrir arquivo PGC para processamento unificado: {e}")
        return [], [], []

    aba_nome = f"PGC {numero_pgc}"
    aba = wb[aba_nome] if aba_nome in wb.sheetnames else wb.active
    ws = aba

    CAMINHO_EMPRESAS = r"C:\PGC\envio_rendimentos\arquivos_gerados\EMPRESAS_NOMECURTO_CNPJ.xlsx"

    def _normalize_key(s):
        import unicodedata, re
        if s is None:
            return ""
        s = str(s).strip().lower()
        s = unicodedata.normalize('NFKD', s)
        s = s.encode('ascii', 'ignore').decode('utf-8')
        s = re.sub(r"[^a-z0-9]+", ' ', s).strip()
        return s

    # carrega mapa empresa -> cnpj
    try:
        df_empresas = pd.read_excel(CAMINHO_EMPRESAS)
        df_empresas.columns = [str(c).strip() for c in df_empresas.columns]
        col_nome = None
        col_cnpj = None
        for c in df_empresas.columns:
            lc = c.lower()
            if 'nome' in lc and ('curto' in lc or 'curt' in lc):
                col_nome = c
            if 'cnpj' in lc:
                col_cnpj = c
        if col_nome and col_cnpj:
            empresas_map = { _normalize_key(r[col_nome]): str(r[col_cnpj]).strip() for _, r in df_empresas.iterrows() }
        else:
            empresas_map = {}
    except Exception:
        empresas_map = {}

    registros_minimos = []
    registros_retencao = []
    registros_cestas = []

    for row in ws.iter_rows(min_row=8):
        # Garantir acesso seguro aos índices
        def val(idx):
            try:
                return row[idx].value if len(row) > idx else None
            except Exception:
                return None

        credor = val(32)  # AG
        desconto_retenc = val(33)  # AH
        empresa_retenc = val(34)  # AI
        desconto_cesta = val(35)  # AJ
        empresa_cesta = val(36)  # AK
        minimo = val(40)  # AO
        empresa_emissao = val(41)  # AP

        # MÍNIMO: somente se empresa_emissao estiver preenchida
        if empresa_emissao and str(empresa_emissao).strip():
            cnpj_lookup = None
            try:
                chave = _normalize_key(empresa_emissao)
                cnpj_lookup = empresas_map.get(chave) if empresas_map else None
            except Exception:
                cnpj_lookup = None

            registros_minimos.append({
                'credor': str(credor).strip() if credor else None,
                'minimo': minimo,
                'empresa': str(empresa_emissao).strip() if empresa_emissao else None,
                'cnpj': cnpj_lookup if cnpj_lookup else None
            })

        # DESCONTOS - retenção clientes
        if desconto_retenc is not None and str(desconto_retenc).strip() != '':
            registros_retencao.append({
                'credor': str(credor).strip() if credor else None,
                'valor': desconto_retenc,
                'empresa': str(empresa_retenc).strip() if empresa_retenc else None
            })

        # DESCONTOS - cestas básicas
        if desconto_cesta is not None and str(desconto_cesta).strip() != '':
            registros_cestas.append({
                'credor': str(credor).strip() if credor else None,
                'valor': desconto_cesta,
                'empresa': str(empresa_cesta).strip() if empresa_cesta else None
            })

    # Salva MINIMO.xlsx se solicitado (comportamento compatível com gerar_minimos_por_coluna_ap)
    if salvar_minimo and registros_minimos:
        df_minimos = pd.DataFrame(registros_minimos)
        if not pasta_saida:
            pasta_saida = os.path.dirname(caminho_arquivo)
        caminho_minimo = os.path.join(pasta_saida, 'MINIMO.xlsx')
        try:
            df_minimos.to_excel(caminho_minimo, index=False)
            logger.info(f"[MÍNIMO] Arquivo MINIMO.xlsx gerado em: {caminho_minimo}")
        except Exception as e:
            logger.error(f"[MÍNIMO] Falha ao salvar MINIMO.xlsx (unificado): {e}")

    # Salvar arquivos de descontos é opcional para manter compatibilidade
    if salvar_descontos and pasta_saida:
        try:
            if registros_retencao:
                pd.DataFrame(registros_retencao).to_excel(os.path.join(pasta_saida, 'DESCONTOS_RETENCAO.xlsx'), index=False)
            if registros_cestas:
                pd.DataFrame(registros_cestas).to_excel(os.path.join(pasta_saida, 'DESCONTOS_CESTAS.xlsx'), index=False)
        except Exception as e:
            logger.warning(f"[DESCONTOS] Falha ao salvar arquivos de descontos: {e}")

    logger.info(f"[MINIMO_UNIFICADO] registros_minimos={len(registros_minimos)}, registros_retencao={len(registros_retencao)}, registros_cestas={len(registros_cestas)}")

    return registros_minimos, registros_retencao, registros_cestas


def extrair_minimos_robusto(aba_pgcs_df, caminho_arquivo, numero_pgc):
    """
    Tenta extrair os dados de mínimo de forma resiliente:
    1. Primeiro tenta com os títulos normalizados
    2. Se falhar, tenta pelas posições fixas (AG, AO, AP)
    """
    if aba_pgcs_df is not None:
        try:
            df_titulos = normalizar_colunas_com_duas_linhas(aba_pgcs_df.copy())
            logger.info("[MÍNIMO] Extração por título foi bem-sucedida.")
            return extrair_minimos_com_base_em_titulos(df_titulos)
        except Exception as e1:
            logger.warning(f'[MÍNIMO] Falha na extração por título: {e1}')
    else:
        logger.warning("[MÍNIMO] DataFrame da aba PGC está ausente.")

    # Fallback por posição
    try:
        logger.info("[MÍNIMO] Tentando extração por posição fixa.")
        return extrair_minimos_por_coluna_fixa(caminho_arquivo, numero_pgc)
    except Exception as e2:
        logger.error(f'[MÍNIMO] Falha também na extração por posição fixa: {e2}')
        raise Exception(f'Erro ao extrair dados de mínimo: {e2}')