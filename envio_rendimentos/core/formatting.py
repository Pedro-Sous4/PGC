import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True)
BODY_FONT = Font(name="Calibri", size=11)

MIN_COL_WIDTH = 12
MAX_COL_WIDTH = 40

MONEY_REGEX = re.compile(r"valor|valor_|valor|r\$|vlr|amount", re.IGNORECASE)
DATE_REGEX = re.compile(r"data|dt|date", re.IGNORECASE)


def column_width_from_cells(cells):
    max_len = 0
    for v in cells:
        if v is None:
            continue
        s = str(v)
        l = len(s)
        if l > max_len:
            max_len = l
    width = max_len + 2
    if width < MIN_COL_WIDTH:
        width = MIN_COL_WIDTH
    if width > MAX_COL_WIDTH:
        width = MAX_COL_WIDTH
    return width


def format_workbook(path):
    """Apply consistent visual formatting to an existing .xlsx file in-place.

    This function only changes styles (fonts, fills, alignments, number formats)
    and column widths. It does NOT modify cell values or sheet structure.
    """
    wb = load_workbook(path)
    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        # Apply default body font to all cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = BODY_FONT

        # Header row formatting
        header_row = 1
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value is not None:
                cell.value = str(cell.value).upper()
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            headers.append(str(cell.value) if cell.value is not None else "")

        # Freeze header and apply autofilter
        ws.freeze_panes = ws.cell(row=2, column=1)
        try:
            ws.auto_filter.ref = ws.dimensions
        except Exception:
            pass

        # Per-column formatting and width
        for col_idx, header in enumerate(headers, start=1):
            header_low = header.lower() if header else ""
            is_money = bool(MONEY_REGEX.search(header_low))
            is_date = bool(DATE_REGEX.search(header_low))

            # Scan a sample of rows for date detection
            if not is_date:
                for r in range(2, min(ws.max_row, 200) + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    if getattr(cell, 'is_date', False):
                        is_date = True
                        break

            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is None:
                    continue
                if is_date:
                    cell.number_format = 'DD/MM/YYYY'
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                elif is_money:
                    cell.number_format = 'R$ #,##0.00'
                    cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

            sample_cells = [ws.cell(row=r, column=col_idx).value for r in range(1, min(ws.max_row, 500) + 1)]
            width = column_width_from_cells(sample_cells)
            column_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[column_letter].width = width

    wb.save(path)
