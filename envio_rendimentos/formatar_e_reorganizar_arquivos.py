import os
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, numbers

# Caminhos base
ORIGEM_BASE = r'C:\PGC\envio_rendimentos\arquivos_gerados\PGC\28'
DESTINO_BASE = r'C:\PGC\envio_rendimentos\arquivos_gerados\PGC\28_ret'

# Colunas com nomes que indicam valores monetários
COLUNAS_MONETARIAS = ['VALOR', 'TOTAL', 'COMISSÃO', 'RECEBER', 'RECEBIDO']

def formatar_planilha_xlsx(caminho):
    wb = load_workbook(caminho)
    for ws in wb.worksheets:
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            coluna_nome = column_cells[0].value
            formatar_como_moeda = False

            if coluna_nome and isinstance(coluna_nome, str):
                for termo in COLUNAS_MONETARIAS:
                    if termo.lower() in coluna_nome.lower():
                        formatar_como_moeda = True
                        break

            for cell in column_cells:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if isinstance(cell.value, (int, float)) and formatar_como_moeda:
                        cell.number_format = 'R$ #,##0.00'
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    wb.save(caminho)

def copiar_e_formatar_arquivos():
    if not os.path.exists(DESTINO_BASE):
        os.makedirs(DESTINO_BASE)

    for credor_nome in os.listdir(ORIGEM_BASE):
        caminho_credor = os.path.join(ORIGEM_BASE, credor_nome)
        if os.path.isdir(caminho_credor):
            destino_credor = os.path.join(DESTINO_BASE, credor_nome)
            os.makedirs(destino_credor, exist_ok=True)

            for arquivo in os.listdir(caminho_credor):
                if arquivo.endswith('.xlsx'):
                    origem_arquivo = os.path.join(caminho_credor, arquivo)
                    destino_arquivo = os.path.join(destino_credor, arquivo)
                    shutil.copy2(origem_arquivo, destino_arquivo)
                    formatar_planilha_xlsx(destino_arquivo)
                    print(f"✔️ {arquivo} copiado e formatado para {destino_credor}")

if __name__ == '__main__':
    copiar_e_formatar_arquivos()
    print("✅ Todos os arquivos foram copiados e formatados com sucesso.")
